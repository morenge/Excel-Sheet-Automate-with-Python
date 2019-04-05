# Python program to read an excel file 
  
# import openpyxl module 
import openpyxl 
  
# Give the location of the file 
path = "C:\\Users\\ginny\\Desktop\\Template\\sample.xlsx"
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path)

#Get workbook active sheet object
#from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have row, column,  
# and coordinate attributes that provide 
# location information for the cell. 
  
# Note: The first row or  
# column integer is 1, not 0. 
  
# Cell object is created by using  
# sheet object's cell() method. 
cell_obj = sheet_obj.cell(row = 1, column = 1) 
  
# Print value of cell object  
# using the value attribute 
print(cell_obj.value) 


#Determine the number of rows

#to open the workbook
#workbook object is created
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

#Print the total number of rows
print(sheet_obj.max_row)
#Print the total number of column
print(sheet_obj.max_column)

#workbook object is created
wb_obj = openpyxl.load_workbook(path)

sheet_obj = wb_obj.active
max_col= sheet_obj.max_column

#Loop will print all columns name
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 1, column = i)
    print(cell_obj.value)

#Will print a particular row value
for i in range(1, max_col + 1):
    cell_obj = sheet_obj.cell(row = 2, column = i)
    print(cell_obj.value, end = "")    


